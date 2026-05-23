from pathlib import Path

import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill,
    Font,
    Alignment,
    Border,
    Side
)
from openpyxl.utils import get_column_letter


# =========================================================
# STYLES LPO
# =========================================================

HEADER_FILL = PatternFill(
    start_color="0088cc",
    end_color="0088cc",
    fill_type="solid"
)

HEADER_FONT = Font(
    color="FFFFFF",
    bold=True
)

THIN_BORDER = Border(
    bottom=Side(
        border_style="thin",
        color="CCCCCC"
    )
)

ITALIC_FONT = Font(
    italic=True,
    color="606060"
)

CENTER = Alignment(
    horizontal="center",
    vertical="center"
)

STATUS_COLORS = {
    "EX": ("000000", "FFFFFF"),
    "EW": ("3d1851", "FFFFFF"),
    "RE": ("5b1a62", "FFFFFF"),
    "CR": ("d20019", "FFFFFF"),
    "EN": ("fabf00", "000000"),
    "VU": ("ffed00", "000000"),
    "NT": ("faf2c7", "000000"),
    "LC": ("78b747", "000000"),
    "DD": ("d4d4d4", "000000"),
}


# =========================================================
# EXPORT EXCEL
# =========================================================

def export_species_excel(
    data,
    output_path
):

    print("Export Excel espèces...")

    df = pd.DataFrame(data)

    wb = Workbook()
    ws = wb.active

    ws.title = "Espèces"

    # =====================================================
    # ECRITURE HEADER
    # =====================================================

    headers = list(df.columns)

    ws.append(headers)

    for cell in ws[1]:

        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # =====================================================
    # ECRITURE DATA
    # =====================================================

    for row in df.itertuples(index=False):

        ws.append(list(row))

    # =====================================================
    # STYLES COLONNES
    # =====================================================

    status_fields = [
        "lr_france",
        "lr_aura",
        "lr_ra",
        "lr_auv"
    ]

    scientific_fields = [
        "lb_nom",
        "sci_name"
    ]

    # mapping nom colonne -> index
    col_index = {
        col: idx + 1
        for idx, col in enumerate(headers)
    }

    # -----------------------------------------------------
    # noms scientifiques
    # -----------------------------------------------------

    for field in scientific_fields:

        if field not in col_index:
            continue

        col = col_index[field]

        for row in range(2, ws.max_row + 1):

            cell = ws.cell(row=row, column=col)

            cell.font = ITALIC_FONT

    # -----------------------------------------------------
    # statuts LR
    # -----------------------------------------------------

    for field in status_fields:

        if field not in col_index:
            continue

        col = col_index[field]

        for row in range(2, ws.max_row + 1):

            cell = ws.cell(row=row, column=col)

            value = str(cell.value)

            if value in STATUS_COLORS:

                bg, fg = STATUS_COLORS[value]

                cell.fill = PatternFill(
                    start_color=bg,
                    end_color=bg,
                    fill_type="solid"
                )

                cell.font = Font(
                    color=fg,
                    bold=True
                )

                cell.alignment = CENTER

    # =====================================================
    # LARGEUR AUTO
    # =====================================================

    for column_cells in ws.columns:

        length = max(
            len(str(cell.value or ""))
            for cell in column_cells
        )

        adjusted = min(length + 4, 40)

        ws.column_dimensions[
            get_column_letter(column_cells[0].column)
        ].width = adjusted

    # =====================================================
    # OPTIONS EXCEL
    # =====================================================

    ws.freeze_panes = "A2"

    ws.auto_filter.ref = ws.dimensions

    # =====================================================
    # SAVE
    # =====================================================

    output_path = Path(output_path)

    output_path.parent.mkdir(
        parents=True,
        exist_ok=True
    )

    wb.save(output_path)

    print(f"Excel exporté : {output_path}")